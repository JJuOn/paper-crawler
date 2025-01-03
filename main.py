import argparse
import json
import sys
import openpyxl
import requests
import os
import csv

from bs4 import BeautifulSoup
from tqdm import tqdm
from typing import Dict, List

def process_conferences(conferences: List[str]) -> List[str]:
    temp = []

    for c in args.conference:
        temp += c.split(",")
    
    if "all" in temp:
        temp = list(conference.keys())

    return temp

def process_years(years: List[str]) -> List[int]:
    temp = []
    for year in years:
        temp += year.split(",")
    
    temp2 = []
    for i in range(len(temp)):
        if "-" in temp[i]:
            start, end = temp[i].split("-")
            temp2 += [str(i) for i in range(int(start), int(end) + 1)]
        elif "~" in temp[i]:
            start, end = temp[i].split("~")
            temp2 += [str(i) for i in range(int(start), int(end) + 1)]
        else:
            temp2.append(temp[i])
    
    return sorted([int(i) for i in list(set(temp2))])

def upper_title(title: str) -> str:
    if title.lower() == "neurips":
        return "NeurIPS"
    else:
        return title.upper()

def get_eccv(year       : int,
             keywords   : List[str]
            ) -> Dict:
    if year == 2024:
        parsed = {"conference": f"ECCV {year}", "papers": [], "authors": []}
        with open("data_eccv2024.csv", "rt") as f:
            reader = csv.reader(f)
            for row in tqdm(reader):
                found = False
                for keyword in keywords:
                    if keyword.lower() in row[0].lower():
                        found = True
                        break
                if found:
                    parsed["papers"].append(row[0])
                    parsed["authors"].append([author.replace("*", "") for author in row[1].split("; ")])


    else:
        res = requests.get("https://www.ecva.net/papers.php")
        soup = BeautifulSoup(res.text, "html.parser")

        papers = soup.findAll("dt", {"class": "ptitle"})
        authors = soup.findAll("dd")
        assert len(papers) == len(authors) // 2

        indices = []
        parsed = {"conference": f"ECCV {year}", "papers": [], "authors": []}
        for i, paper in tqdm(enumerate(papers)):
            if f"eccv_{year}" in paper.find("a")["href"]:
                for keyword in keywords:
                    if keyword.lower() in paper.text or keyword.upper() in paper.text or keyword.capitalize() in paper.text:
                        parsed["papers"].append(paper.text.strip())
                        indices.append(i)
                        break
            else:
                continue
                    
        for idx in indices:
            # parsed["authors"]
            if year in [2020, 2022]:
                author = authors[idx * 2].text.split(",")
                for j in range(len(author)):
                    author[j] = author[j].strip()
                parsed["authors"].append(author)
            elif year in [2018]:
                author = authors[idx * 2].text.split("and")
                for j in range(len(author)):
                    author[j] = author[j].strip().split(',')
                    author[j][0], author[j][1] = author[j][1].strip(), author[j][0].strip()
                    author[j] = " ".join(author[j])
                parsed["authors"].append(author)
            else:
                raise NotImplementedError
    return parsed

def get_neurips(year     : int,
                keywords : List[str]
                ) -> Dict:
    if year == 2023 or year == 2024:
        parsed_oral = {"conference": f"NeurIPS {year} Oral", "papers": [], "authors": []}
        parsed_spotlight = {"conference": f"NeurIPS {year} Spotlight", "papers": [], "authors": []}
        parsed_poster = {"conference": f"NeurIPS {year} Poster", "papers": [], "authors": []}
        res = requests.get(f"https://api2.openreview.net/notes?content.venue=NeurIPS%20{year}%20oral&details=replyCount%2Cpresentation&domain=NeurIPS.cc%2F2023%2FConference&limit=25&offset=0")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 100
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api2.openreview.net/notes?content.venue=NeurIPS%20{year}%20oral&details=replyCount%2Cpresentation&domain=NeurIPS.cc%2F2023%2FConference&limit={limit}&offset={offset}")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]["value"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_oral["papers"].append(title)
                        parsed_oral["authors"].append(row["content"]["authors"]["value"])
                    pbar.update(1)
                offset += limit
        
        res = requests.get(f"https://api2.openreview.net/notes?content.venue=NeurIPS%20{year}%20spotlight&details=replyCount%2Cpresentation&domain=NeurIPS.cc%2F2023%2FConference&limit=25&offset=0")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 400
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api2.openreview.net/notes?content.venue=NeurIPS%20{year}%20spotlight&details=replyCount%2Cpresentation&domain=NeurIPS.cc%2F2023%2FConference&limit={limit}&offset={offset}")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]["value"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_spotlight["papers"].append(title)
                        parsed_spotlight["authors"].append(row["content"]["authors"]["value"])
                    pbar.update(1)
                offset += limit
        
        res = requests.get(f"https://api2.openreview.net/notes?content.venue=NeurIPS%20{year}%20poster&details=replyCount%2Cpresentation&domain=NeurIPS.cc%2F2023%2FConference&limit=25&offset=0")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 1000
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api2.openreview.net/notes?content.venue=NeurIPS%20{year}%20poster&details=replyCount%2Cpresentation&domain=NeurIPS.cc%2F2023%2FConference&limit={limit}&offset={offset}")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]["value"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_poster["papers"].append(title)
                        parsed_poster["authors"].append(row["content"]["authors"]["value"])
                    pbar.update(1)
                offset += limit
        
        return parsed_oral, parsed_spotlight, parsed_poster
    
    elif year == 2022:
        parsed = {"conference": f"NeurIPS {year}", "papers": [], "authors": []}
        res = requests.get(f"https://api.openreview.net/notes?content.venue=NeurIPS+{year}+Accept&details=replyCount&offset=0&limit=1000&invitation=NeurIPS.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 1000
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api.openreview.net/notes?content.venue=NeurIPS+{year}+Accept&details=replyCount&offset={offset}&limit={limit}&invitation=NeurIPS.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
                res_json = json.loads(res.text)
                for note in res_json["notes"]:
                    title = note["content"]["title"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    
                    if keyword_found:
                        parsed["papers"].append(title)
                        parsed["authors"].append(note["content"]["authors"])
                    pbar.update(1)
                offset += limit
        
        return parsed

        
    else:
        parsed = {"conference": f"NeurIPS {year}", "papers": [], "authors": []}
        res = requests.get(f"https://papers.nips.cc/paper/{year}")
        soup = BeautifulSoup(res.text, "html.parser")
        for i, paper in tqdm(enumerate(soup.find_all("div", {"class":"container-fluid"})[0].findAll("li"))):
            title = paper.findAll("a")
            authors = paper.findAll("i")
            keyword_found = False
            for keyword in keywords:
                if keyword.lower() in title[0].text or keyword.upper() in title[0].text or keyword.capitalize() in title[0].text:
                    keyword_found = True
                    break
            if keyword_found:
                parsed["papers"].append(title[0].text)
                parsed["authors"].append(authors[0].text.split(", "))
    return parsed

def get_iclr(year           : int,
             keywords       : List[str],
             ) -> Dict:
    if year == 2024:
        parsed_poster = {"conference": f"ICLR {year}", "papers": [], "authors": []}
        parsed_spotlight = {"conference": f"ICLR {year} spotlight", "papers": [], "authors": []}
        parsed_oral = {"conference": f"ICLR {year} oral", "papers": [], "authors": []}
        # poster session
        res = requests.get(f"https://api2.openreview.net/notes?content.venue=ICLR%20{year}%20poster&details=replyCount%2Cpresentation&domain=ICLR.cc%2F{year}%2FConference&limit=25&offset=0")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 1000
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api2.openreview.net/notes?content.venue=ICLR%20{year}%20poster&details=replyCount%2Cpresentation&domain=ICLR.cc%2F{year}%2FConference&limit={limit}&offset={offset}")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]["value"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_poster["papers"].append(row["content"]["title"]["value"])
                        parsed_poster["authors"].append(row["content"]["authors"]["value"])
                        # parsed_poster["authors"].append(["Anonymous"])
                    pbar.update(1)
                offset += limit
        # spotlight session
        res = requests.get(f"https://api2.openreview.net/notes?content.venue=ICLR%20{year}%20spotlight&details=replyCount%2Cpresentation&domain=ICLR.cc%2F{year}%2FConference&limit=25&offset=0")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 300
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api2.openreview.net/notes?content.venue=ICLR%20{year}%20spotlight&details=replyCount%2Cpresentation&domain=ICLR.cc%2F{year}%2FConference&limit={limit}&offset={offset}")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]["value"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_spotlight["papers"].append(row["content"]["title"]["value"])
                        parsed_spotlight["authors"].append(row["content"]["authors"]["value"])
                        # parsed_spotlight["authors"].append(["Anonymous"])
                    pbar.update(1)
                offset += limit
        # oral session
        res = requests.get(f"https://api2.openreview.net/notes?content.venue=ICLR%20{year}%20oral&details=replyCount%2Cpresentation&domain=ICLR.cc%2F{year}%2FConference&limit=25&offset=0")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 50
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api2.openreview.net/notes?content.venue=ICLR%20{year}%20oral&details=replyCount%2Cpresentation&domain=ICLR.cc%2F{year}%2FConference&limit={limit}&offset={offset}")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]["value"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title or keyword in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_oral["papers"].append(row["content"]["title"]["value"])
                        parsed_oral["authors"].append(row["content"]["authors"]["value"])
                        # parsed_oral["authors"].append(["Anonymous"])
                    pbar.update(1)
                offset += limit
        return parsed_oral, parsed_spotlight, parsed_poster
    elif year == 2023:
        parsed_5 = {"conference": f"ICLR {year} top 5%", "papers": [], "authors": []}
        parsed_25 = {"conference": f"ICLR {year} top 25%", "papers": [], "authors": []}
        parsed_poster = {"conference": f"ICLR {year}", "papers": [], "authors": []}
        # poster session
        res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+poster&details=replyCount&offset=0&limit=25&invitation=ICLR.cc%2F2023%2FConference%2F-%2FBlind_Submission")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 1000
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+poster&details=replyCount&offset={offset}&limit={limit}&invitation=ICLR.cc%2F2023%2FConference%2F-%2FBlind_Submission")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_poster["papers"].append(row["content"]["title"])
                        parsed_poster["authors"].append(row["content"]["authors"])
                    pbar.update(1)
                offset += limit
        # top 25% session
        res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+notable+top+25%25&details=replyCount&offset=0&limit=25&invitation=ICLR.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 1000
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+notable+top+25%25&details=replyCount&offset={offset}&limit={limit}&invitation=ICLR.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_25["papers"].append(row["content"]["title"])
                        parsed_25["authors"].append(row["content"]["authors"])
                    pbar.update(1)
                offset += limit
        # top 5% session
        res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+notable+top+5%25&details=replyCount&offset=0&limit=25&invitation=ICLR.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 1000
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+notable+top+5%25&details=replyCount&offset={offset}&limit={limit}&invitation=ICLR.cc%2F2023%2FConference%2F-%2FBlind_Submission")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_5["papers"].append(row["content"]["title"])
                        parsed_5["authors"].append(row["content"]["authors"])
                    pbar.update(1)
                offset += limit
        return parsed_poster, parsed_25, parsed_5
    else:
        parsed_poster = {"conference": f"ICLR {year}", "papers": [], "authors": []}
        parsed_spotlight = {"conference": f"ICLR {year} spotlight", "papers": [], "authors": []}
        parsed_oral = {"conference": f"ICLR {year} oral", "papers": [], "authors": []}
        # poster session
        res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+Poster&details=replyCount&offset=0&limit=1000&invitation=ICLR.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 1000
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+Poster&details=replyCount&offset=0&limit=1000&invitation=ICLR.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_poster["papers"].append(row["content"]["title"])
                        parsed_poster["authors"].append(row["content"]["authors"])
                    pbar.update(1)
                offset += limit
        # spotlight session
        res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+Spotlight&details=replyCount&offset=0&limit=1000&invitation=ICLR.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 1000
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+Spotlight&details=replyCount&offset=0&limit=1000&invitation=ICLR.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_spotlight["papers"].append(row["content"]["title"])
                        parsed_spotlight["authors"].append(row["content"]["authors"])
                    pbar.update(1)
                offset += limit
        # oral session
        res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+Oral&details=replyCount&offset=0&limit=1000&invitation=ICLR.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
        res_json = json.loads(res.text)
        max_count = res_json["count"]
        offset = 0
        limit = 1000
        with tqdm(range(max_count)) as pbar:
            while offset <= max_count:
                res = requests.get(f"https://api.openreview.net/notes?content.venue=ICLR+{year}+Oral&details=replyCount&offset=0&limit=1000&invitation=ICLR.cc%2F{year}%2FConference%2F-%2FBlind_Submission")
                res_json = json.loads(res.text)
                for row in res_json["notes"]:
                    title = row["content"]["title"]
                    keyword_found = False
                    for keyword in keywords:
                        if keyword.lower() in title or keyword.upper() in title or keyword.capitalize() in title:
                            keyword_found = True
                            break
                    if keyword_found:
                        parsed_oral["papers"].append(row["content"]["title"])
                        parsed_oral["authors"].append(row["content"]["authors"])
                    pbar.update(1)
                offset += limit
        return parsed_poster, parsed_spotlight, parsed_oral


def get_cvpr(year       : int,
             keywords   : List[str]
            ) -> Dict:
    parsed = {"conference": f"CVPR {year}", "papers": [], "authors": []}
    if year == 2024:
        res = requests.get("https://cvpr.thecvf.com/Conferences/2024/AcceptedPapers")
        soup = BeautifulSoup(res.text, "html.parser")
        papers = soup.find_all("tr")[2:-2]
        for i, paper in enumerate(papers):
            title = paper.find("strong")
            if title is None:
                title = paper.find("a")
            title = title.text.strip()
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    parsed["papers"].append(title)
                    authors = paper.find("i").text.strip()
                    authors = authors.split("·")
                    authors = [author.split("(")[0].strip() for author in authors]
                    parsed["authors"].append(authors)
                    break


    elif year in [2021, 2022, 2023]:
        res = requests.get(f"https://openaccess.thecvf.com/CVPR{year}?day=all")
        soup = BeautifulSoup(res.text, "html.parser")
        papers = soup.find_all("dt", {"class": "ptitle"})
        authors = soup.find_all("dd")[1:]
        for i, paper in enumerate(tqdm(papers)):
            for keyword in keywords:
                if keyword.lower() in paper.text.lower():
                    parsed["papers"].append(paper.text)
                    parsed["authors"].append(authors[i * 2].text.split(","))
                    break

    elif year in [2018, 2019, 2020]:
        dates = {2018: ["2018-06-19", "2018-06-20", "2018-06-21"],
                 2019: ["2019-06-18", "2019-06-19", "2019-06-20"],
                 2020: ["2020-06-16", "2020-06-17", "2020-06-18"]}
        for date in dates[year]:
            res = requests.get(f"https://openaccess.thecvf.com/CVPR{year}?day={date}")
            soup = BeautifulSoup(res.text, "html.parser")
            papers = soup.find_all("dt", {"class": "ptitle"})
            authors = soup.find_all("dd")[1:]
            for i, paper in enumerate(papers):
                for keyword in keywords:
                    if keyword.lower() in paper.text.lower():
                        parsed["papers"].append(paper.text)
                        parsed["authors"].append(authors[i * 2].text.split(","))
                        break

    else:
        res = requests.get(f"https://openaccess.thecvf.com/CVPR{year}")
        soup = BeautifulSoup(res.text, "html.parser")
        papers = soup.find_all("dt", {"class": "ptitle"})
        authors = soup.find_all("dd")[1:]
        for i, paper in enumerate(papers):
            for keyword in keywords:
                if keyword.lower() in paper.text.lower():
                    parsed["papers"].append(paper.text)
                    parsed["authors"].append(authors[i * 2].text.split(","))
                    break
    return parsed

def get_iccv(year       : int,
             keywords   : List[str]
            ) -> Dict:
    parsed = {"conference": f"ICCV {year}", "papers": [], "authors": []}

    if year in [2021, 2023]:
        res = requests.get(f"https://openaccess.thecvf.com/ICCV{year}?day=all")
        soup = BeautifulSoup(res.text, "html.parser")
        papers = soup.find_all("dt", {"class": "ptitle"})
        authors = soup.find_all("dd")[1:]
        for i, paper in enumerate(tqdm(papers)):
            for keyword in keywords:
                if keyword.lower() in paper.text.lower():
                    parsed["papers"].append(paper.text)
                    parsed["authors"].append(authors[i * 2].text.split(","))
                    break

    elif year in [2019]:
        dates = {2019: ["2019-10-29", "2019-10-30", "2019-10-31", "2019-11-01"]}
        for date in dates[year]:
            res = requests.get(f"https://openaccess.thecvf.com/ICCV{year}?day={date}")
            soup = BeautifulSoup(res.text, "html.parser")
            papers = soup.find_all("dt", {"class": "ptitle"})
            authors = soup.find_all("dd")[1:]
            for i, paper in enumerate(papers):
                for keyword in keywords:
                    if keyword.lower() in paper.text.lower():
                        parsed["papers"].append(paper.text)
                        parsed["authors"].append(authors[i * 2].text.split(","))
                        break

    else:
        res = requests.get(f"https://openaccess.thecvf.com/ICCV{year}")
        soup = BeautifulSoup(res.text, "html.parser")
        papers = soup.find_all("dt", {"class": "ptitle"})
        authors = soup.find_all("dd")
        for i, paper in enumerate(papers):
            for keyword in keywords:
                if keyword.lower() in paper.text.lower():
                    parsed["papers"].append(paper.text)
                    parsed["authors"].append(authors[i * 2].text.split(","))
                    break
    return parsed

def get_icml(year       : int,
             keywords   : List[str]
            ) -> Dict:
    parsed = {"conference": f"ICML {year}", "papers": [], "authors": []}
    if year == 2024:
        parsed_poster = {"conference": f"ICML {year} Poster", "papers": [], "authors": []}
        parsed_oral = {"conference": f"ICML {year} Oral", "papers": [], "authors": []}
        res = requests.get("https://icml.cc/static/virtual/data/icml-2024-orals-posters.json")
        data = json.loads(res.text)
        for i in tqdm(range(data['count'])):
            title = data['results'][i]['name']
            authors = [author['fullname'] for author in data['results'][i]['authors']]
            eventtype = data['results'][i]['eventtype']
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    if eventtype == 'Poster':
                        parsed_poster["papers"].append(title)
                        parsed_poster["authors"].append(authors)
                        break
                    elif eventtype == 'Oral':
                        parsed_oral["papers"].append(title)
                        parsed_oral["authors"].append(authors)
                        break
        return parsed_poster, parsed_oral
    elif year == 2023:
        parsed_poster = {"conference": f"ICML {year} Poster", "papers": [], "authors": []}
        parsed_oral = {"conference": f"ICML {year} Oral", "papers": [], "authors": []}
        res = requests.get("https://icml.cc/static/virtual/data/icml-2023-orals-posters.json")
        data = json.loads(res.text)
        for i in tqdm(range(data['count'])):
            title = data['results'][i]['name']
            authors = [author['fullname'] for author in data['results'][i]['authors']]
            eventtype = data['results'][i]['eventtype']
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    if eventtype == 'Poster':
                        parsed_poster["papers"].append(title)
                        parsed_poster["authors"].append(authors)
                        break
                    elif eventtype == 'Oral':
                        parsed_oral["papers"].append(title)
                        parsed_oral["authors"].append(authors)
                        break
        return parsed_poster, parsed_oral
    elif year < 2023:
        res = requests.get("https://proceedings.mlr.press")
        soup = BeautifulSoup(res.text, "html.parser")
        proceedings_list = soup.find_all("ul", {"class": "proceedings-list"})[1].find_all("li")
        for proceeding in proceedings_list:
            if year >= 2017:
                if f"Proceedings of ICML {year}" in proceeding.text and "Workshop" not in proceeding.text:
                    href = proceeding.find("a")["href"]
                    break
            else:
                if f"ICML {year} Proceedings" in proceeding.text and "Workshop" not in proceeding.text:
                    href = proceeding.find("a")["href"]
                    break

        res = requests.get(f"https://proceedings.mlr.press/{href}")
        soup = BeautifulSoup(res.text, "html.parser")
        papers = soup.find_all("div", {"class": "paper"})
        for paper in tqdm(papers):
            title = paper.find("p", {"class": "title"}).text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = paper.find("p", {"class": "details"}).find("span", {"class": "authors"}).text.split(",")
                    authors = [a.strip() for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
        return parsed

def get_acl(year    : int,
            keywords: List[str]) -> Dict:
    parsed = {"conference": f"ACL {year}", "papers": [], "authors": []}
    if year >= 2021:
        res = requests.get(f"https://aclanthology.org/events/acl-{year}/")
        soup = BeautifulSoup(res.text, "html.parser")
        long_div = soup.find_all("div", {"id": f"{year}acl-long"})[0]
        papers = long_div.find_all("span", {"class": "d-block"})
        for paper in tqdm(papers):
            a_list = paper.find_all("a")
            title = a_list[0].text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = a_list[1:]
                    authors = [a.text for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
                    break
            

        short_div = soup.find_all("div", {"id": f"{year}acl-short"})[0]
        papers = short_div.find_all("span", {"class": "d-block"})
        for paper in tqdm(papers):
            a_list = paper.find_all("a")
            title = a_list[0].text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = a_list[1:]
                    authors = [a.text for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
                    break

    elif year == 2020:
        res = requests.get(f"https://aclanthology.org/events/acl-{year}/")
        soup = BeautifulSoup(res.text, "html.parser")
        div = soup.find_all("div", {"id": f"{year}acl-main"})[0]
        papers = div.find_all("span", {"class": "d-block"})
        for paper in tqdm(papers):
            a_list = paper.find_all("a")
            title = a_list[0].text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = a_list[1:]
                    authors = [a.text for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
                    break

    elif year == 2019:
        res = requests.get(f"https://aclanthology.org/events/acl-{year}/")
        soup = BeautifulSoup(res.text, "html.parser")
        div = soup.find_all("div", {"id": f"p{str(year)[2:]}-1"})[0]
        papers = div.find_all("span", {"class": "d-block"})
        for paper in tqdm(papers):
            a_list = paper.find_all("a")
            title = a_list[0].text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = a_list[1:]
                    authors = [a.text for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
                    break

    elif year == 2018:
        res = requests.get(f"https://aclanthology.org/events/acl-{year}/")
        soup = BeautifulSoup(res.text, "html.parser")
        long_div = soup.find_all("div", {"id": f"p{str(year)[2:]}-1"})[0]
        papers = long_div.find_all("span", {"class": "d-block"})
        for paper in tqdm(papers):
            a_list = paper.find_all("a")
            title = a_list[0].text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = a_list[1:]
                    authors = [a.text for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
                    break
            
        short_div = soup.find_all("div", {"id": f"p{str(year)[2:]}-2"})[0]
        papers = short_div.find_all("span", {"class": "d-block"})
        for paper in tqdm(papers):
            a_list = paper.find_all("a")
            title = a_list[0].text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = a_list[1:]
                    authors = [a.text for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
                    break
    
    else:
        raise
    
    return parsed

def get_emnlp(year    : int,
              keywords: List[str]) -> Dict:
    parsed = {"conference": f"EMNLP {year}", "papers": [], "authors": []}
    if year == 2022:
        res = requests.get("https://preview.aclanthology.org/emnlp-22-ingestion/volumes/2022.emnlp-main/")
        soup = BeautifulSoup(res.text, "html.parser")
        papers = soup.find_all("span", {"class": "d-block"})
        for paper in tqdm(papers):
            a_list = paper.find_all("a")
            title = a_list[0].text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = a_list[1:]
                    authors = [a.text for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
                    break
                
    elif year >= 2020:
        res = requests.get(f"https://aclanthology.org/events/emnlp-{year}/")
        soup = BeautifulSoup(res.text, "html.parser")
        div = soup.find_all("div", {"id": f"{year}emnlp-main"})[0]
        papers = div.find_all("span", {"class": "d-block"})
        for paper in tqdm(papers):
            a_list = paper.find_all("a")
            title = a_list[0].text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = a_list[1:]
                    authors = [a.text for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
                    break
    
    elif year >= 2018:
        res = requests.get(f"https://aclanthology.org/events/emnlp-{year}/")
        soup = BeautifulSoup(res.text, "html.parser")
        div = soup.find_all("div", {"id": f"d{str(year)[2:]}-1"})[0]
        papers = div.find_all("span", {"class": "d-block"})
        for paper in tqdm(papers):
            a_list = paper.find_all("a")
            title = a_list[0].text
            for keyword in keywords:
                if keyword.lower() in title.lower():
                    authors = a_list[1:]
                    authors = [a.text for a in authors]
                    parsed["papers"].append(title)
                    parsed["authors"].append(authors)
                    break
    
    else:
        raise
    
    return parsed

conference = {
    "eccv"    : get_eccv,
    "neurips" : get_neurips,
    "iclr" : get_iclr,
    "cvpr" : get_cvpr,
    "iccv" : get_iccv,
    "icml" : get_icml,
    "acl" : get_acl,
    "emnlp" : get_emnlp,
}

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-c", "--conference", required=True, type=str, nargs="+")
    parser.add_argument("-y", "--year", required=True, type=str, nargs="+") # 
    parser.add_argument("-k", "--keywords", required=True, type=str, nargs="+")

    args = parser.parse_args()

    # Process arguments
    conferences = process_conferences(args.conference)
    years = process_years(args.year)

    parseds = []
    for conf in conferences:
        print(f"Start {conf}")
        for year in years:
            # try:
            parsed = conference[conf](year, args.keywords)
            if parsed is not None:
                parseds.append(parsed)
            # except:
            #     print(f"{conf} {year} failed")
            #     pass

    wb = openpyxl.Workbook()
    offset = 1
    for parsed in parseds:
        if isinstance(parsed, tuple):
            if len(parsed) == 2:
                sheet = wb.worksheets[0]
                for session in range(len(parsed)):
                    for row, (paper, author) in enumerate(zip(parsed[session]["papers"], parsed[session]["authors"]), offset):
                        sheet.cell(row=row, column=1).value = parsed[session]["conference"]
                        sheet.cell(row=row, column=2).value = paper
                        sheet.cell(row=row, column=3).value = author[0]
                    offset += len(parsed[session]["papers"])
            elif len(parsed) == 3: # poster, spotlight, oral
                sheet = wb.worksheets[0]
                for session in range(len(parsed)):
                    for row, (paper, author) in enumerate(zip(parsed[session]["papers"], parsed[session]["authors"]), offset):
                        sheet.cell(row=row, column=1).value = parsed[session]["conference"]
                        sheet.cell(row=row, column=2).value = paper
                        sheet.cell(row=row, column=3).value = author[0]
                    offset += len(parsed[session]["papers"])
        else:
            sheet = wb.worksheets[0]
            for row, (paper, author) in enumerate(zip(parsed["papers"], parsed["authors"]), offset):
                sheet.cell(row=row, column=1).value = parsed["conference"]
                sheet.cell(row=row, column=2).value = paper
                sheet.cell(row=row, column=3).value = author[0]
            offset += len(parsed["papers"])

    if not len(conferences) == len(list(conference.keys())):
        upper_conference = "_".join([upper_title(c) for c in conferences])
    else:
        upper_conference = 'ALL'

    years = "_".join([str(y) for y in years])
    keywords = [k.lower() for k in args.keywords]
    keywords = "_".join(sorted(keywords))

    save_path = f"{upper_conference}_{years}_{keywords}.xlsx"
    wb.save(save_path)
